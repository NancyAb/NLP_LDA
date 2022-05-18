# -*- coding: utf-8 -*-
"""
Created on Wed Jan 27 20:27:06 2021

@author: nabramson
"""

# 
# import data
# 

import os,sys,re,string

if not sys.warnoptions:
    import warnings
    warnings.simplefilter("ignore")
    warnings.filterwarnings("ignore", category=DeprecationWarning)

import random, sqlite3
import pandas as pd
import numpy as np

from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.decomposition import TruncatedSVD #LSA Latent Semantic Indexing 

import pyLDAvis
import pyLDAvis.sklearn

from sklearn.decomposition import LatentDirichletAllocation
import nltk
from nltk.stem import PorterStemmer

from datetime import datetime

import pickle

# MS Excel supporting utilities

from openpyxl.utils import get_column_letter

try:
    sys.path.remove('.') 
    sys.path.append( '.')  
except:
    sys.path.append( '.')
    
base = 'Health_' + datetime.today().strftime('%Y-%m-%d')

def write_data(workbook, final_data_df):
    
    topics = workbook.add_worksheet( 'data')
    
    topics.set_column('A:A', 15)
    topics.set_column('B:B', 25)
    topics.set_column('C:D', 15)
    topics.set_column('E:E', 80)
    topics.set_column('F:J', 15)
    topics.set_column('K:Q', 0)
    
    final_data_df.fillna('', inplace = True)
    
    cols = list(final_data_df.columns)
    
    data = final_data_df.values.tolist()
    
     # convert cols to dictionary with header for use with options
         
    columns  = [{'header':c.replace('_', ' ').title()} for c in cols ]  
    
    # Options to use in the table.
    options = {'data': data,
               'columns':columns }
   
# Add a table to the worksheet.
    last_col = get_column_letter(len(cols))    
    topics.add_table('A1:'  + last_col +  str(len(final_data_df) + 1), options)
    
    return

def write_priors(workbook, topics_df):
    
    topics = workbook.add_worksheet( 'priors')
    
    topics.set_column('A:A', 18)
    topics.set_column('B:B', 10)
    topics.set_column('C:C', 15)
    topics.set_column('D:D', 18)
    topics.set_column('E:E', 12)
    
    topics_df.fillna('', inplace = True)
    
    data = topics_df.values.tolist()

    # Options to use in the table.
    # convert cols to dictionary with header for use with options
         
    columns  = [{'header':c.replace('_', ' ').title()} for c in topics_df.columns ]  
    
    options = {'data': data,
               'columns': columns}

# Add a table to the worksheet.

    last_col = get_column_letter(topics_df.shape[1])  
    topics.add_table('A1:' + last_col + str(len(topics_df) + 1), options)
    
    return

def preptext_health(df, text_col):
    
    df_priors = df.copy()
    clean_col = text_col + '_clean'
    
    df_priors[clean_col] = df_priors[text_col]
    
    stopwords = nltk.corpus.stopwords.words('english') 
    
    # keep negative words
     
    list_negative = [n for n in stopwords if "n't" in n] 
    
    def remove_contractions(text, dict_negative):
        
        l = text.lower()
        for k in dict_negative:
            l = l.replace(k,dict_negative[k] + " not")
        
        return l
    
    def remove_stop(text, stopwords):
        
        l = text
        l = re.split('\\W+',l)
        
        l = [w.lower() for w in l if w.lower() not in stopwords ]
        table = str.maketrans({key: None for key in string.punctuation})
        l = [str(s.translate(table)) for s in l]
        
        return ' '.join(l), len(l)
    
    #df_priors[clean_col] = df_priors[text_col].apply(lambda x: remove_contractions(x, dict_negative))
    
    df_priors[clean_col] = df_priors[clean_col].apply(lambda x: remove_stop(x, stopwords))
    df_priors[[clean_col, 'cleaned_length']] = pd.DataFrame(df_priors[clean_col].tolist(), index=df_priors.index)


    return df_priors

def preptext(df, text_col):
    
    df_priors = df.copy()
    clean_col = text_col + '_clean'
    ps = PorterStemmer()
    
    def clean_list(text, stopwords, ps):
        
        # l= 'summary depo'
        # keep colon from string.punctuation
        # no_colonpunctuation = '!"#$%&\'()*+,-./:;<=>?@[\\]^_`{|}~'

        l = text.replace('-',' ')
        l = l.replace('.mp4" class="video-url"><input type="hidden" name="" value="https://images-na.ssl-images', '')
        l = l.replace('video-block"></div><input type="hidden" name="" value="https://images','')
        l = l.replace('video-block"></div>', '')
        l = l.replace('<div id="video-block-', '')
        l = l.replace('<div id=', '')
        l = l.replace('"','')
        l = l.replace('&amp','')
        l = l.replace('(',' ')
        l = l.replace(')',' ')
        l = l.replace('/',' ')
        l = l.replace(';',' ')
        l = l.replace('k9','dog')
        l = l.replace('canine','dog')
        l = l.replace('feline','cat')
        l = l.replace('1st','first')
        l = l.replace('2nd','second')
        l = l.replace('3rd','third')
        l = l.replace('2x','times')
        l = l.replace('3x','times')
        l = re.sub(r'[0-9]+', '', l)
        l = re.split('\\W+',l)
        l = [w for w in l if (  not w.isdigit() ) ]  # remove all numbers as words
        
        # lemmenitize in spark
        
        l = [ps.stem(w).lower() for w in l if w.lower() not in stopwords ]
        l = [ i.replace('describ','') for i in l ]
        l = [ i.replace('item','')  for i in l ]
        table = str.maketrans({key: None for key in string.punctuation})
        l = [str(s.translate(table)) for s in l]
        
        return ' '.join(l), len(l)
      
    # only keep in baseline that has not been reviewed 
    # keep reviewed for printing off later
    #nltk.download('stopwords')
    more_noise = ['aspacingsm','best', 'aspacingsmal', 'aspacingsmall', 'aspacingtopmini','better', 'classaect',
                  'classasect', 'classasection', 'cmcrarpdrvwtxt', 'cute', 'div', 'easy', 'enjoydiv',
                  'enjoy', 'enjoys', 'enjoyed','hate', 'hates', 'hated',
                  'favorite', 'good', 'great', 'happy', 'idvideoblockRXMUYLQX', 'ie', 'like', 'love',
                  'loves','loved', 'n', 'nice', 'perfect', 'perfectly', 'product', 'quality', 'really',
                  'datahookproductlinklink','classalinknor','last', 'long', 'time',
                  'datahookproductlinklinked' , 'classalinknormal' ,
                  'recommend', 'ref', 'super', 'typehidden', 'typehiddenclassasect', 'videoblockdivinput',
                  'well', 'work']
    
    more_noise.sort()
    print(more_noise)
    stopwords = nltk.corpus.stopwords.words('english') + more_noise
    
    df_priors[clean_col] = df_priors[text_col].apply(lambda x: clean_list(x, stopwords, ps))
    df_priors[[clean_col, 'cleaned_length']] = pd.DataFrame(df_priors[clean_col].tolist(), index=df_priors.index)

    return df_priors

def get_data_from_db(base =  'pets', text_col = 'review_text',
                     select_where = " where 1=1 limit 100 "):
    
    conn = sqlite3.connect(base + '/' + base + ".db")
    query_all = "SELECT * from reviews " + select_where
    query = "SELECT asin, review_text from reviews" + select_where + ";"
    df = pd.read_sql_query(query, conn)
    df_all = pd.read_sql_query(query_all, conn)    
    return df

def get_urls_from_db(base =  'pets', text_col = 'review_text',
                     select_where = " where 1=1 limit 100 "):
    
    conn = sqlite3.connect(base + '/' + base + ".db")
    query_urls = "SELECT * from products " + select_where
    df_urls = pd.read_sql_query(query_urls, conn)    
    return df_urls

def lsa_calc(vectorizer, topics, x, col_names, tfidf_df):
    
    U, S, V = np.linalg.svd(x.toarray(), full_matrices=False)
        
    # inspect shapes of the matrices
    print(U.shape, S.shape, V.shape)
        
    lsa_obj = TruncatedSVD(n_components=topics, n_iter=400, random_state=42)
    tfidf_lsa_data = lsa_obj.fit_transform(tfidf_df)
    Sigma = lsa_obj.singular_values_
    V_T = lsa_obj.components_.T
    
    print(tfidf_df.shape, Sigma.shape, V_T.shape)

    #sns.barplot(x=list(range(len(Sigma))), y = Sigma)
    
    term_topic_matrix = pd.DataFrame(data=V_T, 
                                 index = col_names, 
                                 columns = [f'Latent_concept_{str(r).zfill(3)}' for r in range(0,V_T.shape[1])])
    term_topic_matrix['context_number'] = 'Latent_concept'
    
    
    topics_df = pd.DataFrame(columns= ['context_number', 'context'])
    
    for i in range(topics): #i=69; i= 70
        
        i= str(i).zfill(3)
        print(i)
         
        threshhold = 0.20
        
        #topics_i = term_topic_matrix[[f'Latent_concept_{i}', 'context']] ; ' '.join({'ab', 'cde', 'efg'})
        
        topics_i = term_topic_matrix[term_topic_matrix[f'Latent_concept_{i}']>threshhold][[f'Latent_concept_{i}', 'context_number']]
        topics_i = topics_i.sort_values(by = [f'Latent_concept_{i}'], ascending=False).reset_index()
        topics_i['context_number'] = 'Latent_concept_' + str(i)
         
        topics_i = topics_i.groupby('context_number')['index'].apply(' '.join).reset_index() 
        topics_i.columns = ['context_number', 'context']
        topics_i['context'] = topics_i['context'].apply(lambda x: ' '.join(set(x.split())))     
        topics_df = pd.concat([topics_df, topics_i] , axis=0)
            
        final_data =np.dot( tfidf_df , V_T)
        
        pickle.dump(V_T, open('./Pickle/V_T' +  datetime.today().strftime('%Y-%m-%d') + '.pkl','wb'))
        pickle.dump(topics_df, open('./Pickle/topics_df'  + datetime.today().strftime('%Y-%m-%d') + '.pkl','wb'))
        
        # pad the topic number with 00 so sort is nice using zfill
        
        final_data_df = pd.DataFrame(data=final_data, 
                                     columns = [f'Latent_concept_{str(r).zfill(3)}' for r in range(0,V_T.shape[1])])
        
        final_topic_prob = final_data_df.copy()
    
    final_data =np.dot( tfidf_df , V_T)
    
    pickle.dump(V_T, open('./Pickle/V_T' +  datetime.today().strftime('%Y-%m-%d') + '.pkl','wb'))
    pickle.dump(topics_df, open('./Pickle/topics_df'  + datetime.today().strftime('%Y-%m-%d') + '.pkl','wb'))
            
    return final_data, term_topic_matrix, V_T

def lda_calc(vectorizer, df_pets, column_name, topics):
    
    df = df_pets.copy()
    column = column_name
    
    data = df[column]
    xlda = df[ column_name + '_clean']    
    data_matrix = vectorizer.fit_transform(xlda)
    print(data_matrix.shape)
    lda_model = LatentDirichletAllocation(
        n_components=topics,  # Number of topics
        learning_method='online',
        random_state=20,
        n_jobs=-1  # Use all available CPUs
    )
    lda_output = lda_model.fit_transform(data_matrix)
    data = pyLDAvis.sklearn.prepare(lda_model, data_matrix, vectorizer, mds='tsne')
    return data, lda_output

def main():
    
    text_physical = 'Why or why not?'
    text_mental = 'Why or why not?.1'
    health_file = "mental-heath-in-tech-2016_20161114.csv"
     
    df_all_data = pd.read_csv(health_file)
    
    #col = text_physical
    for col in [text_physical, text_mental]: 
        df_all_data[col].fillna('', inplace = True)
            
        df_health = preptext_health(df_all_data, col)
        
        size = len(df_health)
        filename = "health_why_clean_" + str(int(size/1000)) +  "k.csv"
        
        # try ngram for fun
        
        vectorizer = TfidfVectorizer(ngram_range=(1,3),
                                         max_features=50000,
                                         min_df = 1,
                                         max_df = 1.0)
        
        # words as columns   
        
        x = vectorizer.fit_transform(df_health[col + '_clean'])
        col_names = vectorizer.get_feature_names()
        
        tfidf_df = pd.DataFrame(x.toarray(), 
                            columns=col_names)
            
        topics = 15
    
        data, lda_output = lda_calc(vectorizer, df_health, col, topics)
        
        if col == "Why or why not?":
            filename = "physical_health_" + str(topics) + "_" + str(random.randrange(0, 100) )   + ".html"
        else:
            filename = "mental_health_" + str(topics) + "_" + str(random.randrange(0, 100) )   + ".html"
            
        fp = open(filename,"a")
        pyLDAvis.save_html(data, fp)
        fp.close()
        
        print(df_health[col + '_clean'].value_counts())